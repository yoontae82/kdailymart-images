"""listing-agent

역할: Global SKU 데이터를 조립해 쇼피 대용량업로드 형식의 실제 파일(.xlsx)을 생성한다.
입력: 카피(localization-agent), 이미지(image-agent), 카테고리ID/속성값(rulebook/category_master.json),
      마진(margin-agent), SKU/재고/DTS/치수 등 물류 정보
출력: 사람이 쇼피 셀러센터에 직접 업로드할 .xlsx 파일
반려 조건: 카테고리/DTS/속성값이 category_master.json 허용 범위를 벗어나면 조립 단계에서 즉시 거부

파이프라인 순서: 5단계.

운영 방식 (Phase 1 - 로그인 자동화 없이 시작):
쇼피 셀러센터 로그인 자동화(브라우저자동화/Open API)는 크리덴셜·2FA·승인 절차가 걸려있어
당장 하지 않는다. 대신 이 파일이 정확한 형식으로 만들어지면, 사람이 쇼피 셀러센터에서
"대용량업로드" 파일을 직접 첨부해서 올리고 개별 Publish 버튼을 누르는 방식으로 간다.
mass_upload()/publish_individually()는 이 수동 절차가 병목으로 확인될 경우 고려할 Phase 2
(브라우저자동화)를 위해 남겨둔 스텁이며, 지금 당장 필요하지 않다.

검증된 절차 (운영 룰북 v2):
- 등록 방식: Global SKU -> Mass Upload -> 개별 Publish (개별 Publish가 정답 경로, Mass Publish는
  "No Product Found"로 실패한 이력이 있어 사용하지 않음)
- Global SKU Price 컬럼(K)에는 원가만 입력 (판매가 아님, margin-agent 산출값 사용)
- 고정 컬럼 레이아웃은 category_master.json의 fixed_columns 참고
"""

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from image import ProcessedImageSet
from localization import LocalizedCopy
from margin import MarginResult

Country = Literal["SG", "MY"]

CATEGORY_MASTER_PATH = Path(__file__).resolve().parent.parent / "rulebook" / "category_master.json"


@dataclass
class ListingResult:
    item_id: str | None
    sku: str
    status: Literal["global_sku_only", "file_generated", "mass_uploaded", "live", "publish_failed"]
    failure_reason: str | None = None


def load_category_master(path: Path = CATEGORY_MASTER_PATH) -> dict:
    """rulebook/category_master.json을 로드한다."""
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _find_category(category_master: dict, country: Country, category_id: int) -> dict:
    for entry in category_master["categories"].get(country, []):
        if entry.get("category_id") == category_id:
            return entry
    raise ValueError(
        f"category_master.json에 없는 카테고리입니다 (country={country}, category_id={category_id}). "
        "실제 쇼피 템플릿(HiddenCatProps 등)을 파싱해서 먼저 채워야 합니다."
    )


def build_mass_upload_row(
    copy: LocalizedCopy, images: ProcessedImageSet, margin: MarginResult,
    category_id: int, country: Country, sku: str, stock: int, days_to_ship: int,
    weight_kg: float, dimensions_cm: tuple[float, float, float],
    attributes: dict[str, str], parent_sku: str | None = None, brand: str = "No Brand",
    category_master: dict | None = None,
) -> dict[str, str]:
    """category_master.json의 컬럼 규칙에 맞춰 Mass Upload 1행 데이터를 조립한다.
    반환값은 {컬럼문자: 값} 형태이며 generate_mass_upload_file()에 그대로 넘기면 된다.

    카테고리ID/DTS/속성값이 category_master.json 허용 범위를 벗어나거나 mandatory 속성이
    빠지면 ValueError를 던진다 - 이 검증을 build 단계에서 하는 이유는, 형식이 틀린 파일이 만들어진
    뒤에야(즉 사람이 쇼피에 업로드해본 뒤에야) 오류를 알게 되는 걸 막기 위해서다.

    ⚠️ Global SKU Price(K 컬럼, margin.cost_price_krw)의 통화/공식은 아직 확정이 안 됐다.
    실제 템플릿 가이드 문구는 "Global SKU Price를 RMB로 입력, Shop SKU Price = Global SKU Price *
    Marketplace Exchange Rate * Market Price Adjustment Ratio + Hidden Price"라고 되어 있는데,
    이 프로젝트의 다른 문서(운영 룰북 v2)는 KRW 기준/Adjustment Rate 단순 공식을 전제로 한다.
    실제 값을 넣기 전에 반드시 사람이 쇼피 셀러센터에서 이 통화/공식을 재확인해야 한다 -
    README의 "가격 컬럼(K) 통화/공식 확인 필요" 항목 참고.

    TODO(정식 운영 전 필요 사항):
    - 위 K컬럼 통화/공식 확인
    - Parent SKU(D 컬럼)를 언제 채우고 언제 비워두는지 규칙이 아직 없음 (변형상품이 있는 경우에만
      필요한 것으로 보이나 룰북에 명확한 기준이 없어 추가 확인 필요)
    - Brand(AD 컬럼)는 실제 브랜드ID 목록(Brand 시트)이 없어 기본값 "No Brand"로 두고 있음 -
      브랜드 등록이 필요해지면 실제 Brand ID 목록을 받아 검증 로직 추가
    """
    master = category_master if category_master is not None else load_category_master()
    entry = _find_category(master, country, category_id)

    dts_rule = entry["days_to_ship"]
    if days_to_ship in dts_rule.get("banned", []):
        raise ValueError(
            f"DTS={days_to_ship}는 이 카테고리에서 금지값입니다 (banned={dts_rule['banned']}). {dts_rule.get('note', '')}"
        )
    if not any(lo <= days_to_ship <= hi for lo, hi in dts_rule.get("allowed", [])):
        raise ValueError(f"DTS={days_to_ship}는 허용 범위({dts_rule.get('allowed')}) 밖입니다.")

    row: dict[str, str] = {}
    fixed = master["fixed_columns"]
    values_by_field = {
        "category_id": str(category_id),
        "product_name": copy.title,
        "description": copy.description,
        "parent_sku": parent_sku or "",
        "cost_price_krw": str(margin.cost_price_krw),
        "stock": str(stock),
        "sku": sku,
        "cover_image_url": images.cover_image_url,
        "weight_kg": str(weight_kg),
        "length_cm": str(dimensions_cm[0]),
        "width_cm": str(dimensions_cm[1]),
        "height_cm": str(dimensions_cm[2]),
        "days_to_ship": str(days_to_ship),
        "brand": brand,
    }
    for col, field in fixed.items():
        if col.startswith("$"):
            continue
        if field in values_by_field:
            row[col] = values_by_field[field]
    image_columns = ["O", "P", "Q", "R", "S", "T", "U", "V"]  # Item Image 1~8
    for i, col in enumerate(image_columns):
        if i < len(images.item_image_urls):
            row[col] = images.item_image_urls[i]

    for attr_name, value in attributes.items():
        attr_rule = entry["attributes"].get(attr_name)
        if attr_rule is None:
            raise ValueError(f"'{attr_name}'은 이 카테고리에 정의되지 않은 속성입니다.")
        allowed_values = attr_rule.get("allowed_values")
        if allowed_values is not None and value not in allowed_values:
            raise ValueError(f"'{attr_name}'의 값 '{value}'는 허용값({allowed_values})에 없습니다.")
        row[attr_rule["column"]] = value

    missing_mandatory = [
        name for name, rule in entry["attributes"].items()
        if rule.get("mandatory") and name not in attributes
    ]
    if missing_mandatory:
        raise ValueError(f"이 카테고리의 필수(mandatory) 속성이 빠졌습니다: {missing_mandatory}")

    return row


def _load_workbook_tolerant(template_path: Path):
    """openpyxl로 워크북을 열되, 실제 쇼피 템플릿에서 확인된 손상된 sheetView 속성
    (activePane="bottom_left" - OOXML 표준값 "bottomLeft"가 아니라 openpyxl이 거부함)이
    있으면 자동으로 고쳐서 연다. 데이터는 전혀 건드리지 않고 view 상태값 표기만 정규화한다.
    """
    import openpyxl

    try:
        return openpyxl.load_workbook(template_path)
    except ValueError as e:
        cause_text = f"{e}\n{e.__cause__}"
        if "bottomLeft" not in cause_text and "topLeft" not in cause_text:
            raise
        import re
        import tempfile
        import zipfile

        with tempfile.TemporaryDirectory() as tmp_dir:
            extract_dir = Path(tmp_dir) / "extracted"
            with zipfile.ZipFile(template_path) as zf:
                zf.extractall(extract_dir)

            fixed_any = False
            for sheet_xml in (extract_dir / "xl" / "worksheets").glob("*.xml"):
                text = sheet_xml.read_text(encoding="utf-8")
                new_text = re.sub(r'activePane="[a-zA-Z_]+"', lambda m: (
                    m.group(0).replace("bottom_left", "bottomLeft")
                    .replace("bottom_right", "bottomRight")
                    .replace("top_left", "topLeft")
                    .replace("top_right", "topRight")
                ), text)
                if new_text != text:
                    sheet_xml.write_text(new_text, encoding="utf-8")
                    fixed_any = True
            if not fixed_any:
                raise

            fixed_path = Path(tmp_dir) / "fixed.xlsx"
            with zipfile.ZipFile(fixed_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for file_path in extract_dir.rglob("*"):
                    if file_path.is_file():
                        zf.write(file_path, file_path.relative_to(extract_dir))

            wb = openpyxl.load_workbook(fixed_path)
            # 고친 파일은 임시 디렉토리와 함께 사라지므로, 원본 대신 이 워크북 객체를 그대로 쓴다.
            return wb


def generate_mass_upload_file(
    rows: list[dict[str, str]], output_path: Path, template_path: Path | None = None,
    sheet_name: str = "Template", data_start_row: int = 7,
) -> Path:
    """조립된 행들을 실제 .xlsx 파일로 저장한다. 로그인 자동화 없이 사람이 이 파일을 쇼피
    셀러센터에서 직접 "대용량업로드"로 첨부하는 Phase 1 워크플로우의 산출물이다.

    template_path를 주면 실제 쇼피 대용량업로드 템플릿 파일을 열어 그 위에 값만 채운다
    (권장 - 쇼피가 기대하는 다른 시트(HiddenCatProps 등)와 헤더가 그대로 보존된다). 실제 템플릿은
    "Template" 시트의 1~6행이 내부 식별자/헤더/가이드이고, 7행부터 실제 상품 데이터가 시작된다
    (Shopee_mass_upload_global_sku_20260716_advanced_template.xlsx로 확인됨).
    template_path가 없으면 컬럼 문자만으로 새 파일을 만드는데, 이건 로직 검증용 데모일 뿐이며
    실제 쇼피 시트 구성과 다를 수 있어 이 상태로 바로 업로드하면 안 된다.

    TODO(정식 운영 전 필요 사항):
    - 지금은 매번 template_path를 명시적으로 넘겨야 한다. 실제 운영에 들어가면 이 템플릿 파일을
      rulebook/ 밑에 고정 보관하고 기본값으로 참조하도록 바꿀 것 (템플릿이 쇼피 쪽에서 갱신되면
      다시 받아서 교체해야 함 - 카테고리 트리/속성이 바뀔 수 있음)
    - 여러 행을 이어 붙일 때 기존 "Upload sample" 시트의 예시 행과 겹치지 않는지 확인 필요
    """
    if template_path is not None:
        wb = _load_workbook_tolerant(template_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        start_row = max(ws.max_row + 1, data_start_row)
    else:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DEMO - not a real Shopee template"
        if rows:
            for col in sorted({c for row in rows for c in row}):
                ws[f"{col}1"] = col
        start_row = 2

    for i, row in enumerate(rows):
        for col, value in row.items():
            ws[f"{col}{start_row + i}"] = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def mass_upload(rows: list[dict], country: Country) -> list[ListingResult]:
    """[Phase 2 - 지금 당장 필요하지 않음] 조립된 행들을 쇼피 셀러센터에 자동으로 Mass Upload 한다.
    Phase 1(generate_mass_upload_file + 사람이 수동 업로드)이 병목으로 확인될 때만 고려한다.

    TODO(Phase 2를 실제로 진행하기로 결정한 경우 필요 사항):
    - 브라우저 자동화 방식 결정: (a) Playwright/Selenium으로 셀러센터 UI 직접 조작
      (b) 쇼피 Open API(공식 연동) 사용 가능 여부 확인 후 전환
    - 셀러센터 로그인 크리덴셜 및 2FA/OTP 처리 방식 (자동화 시 OTP 수신 방법)
    - seller.shopee.kr(KRSC) 계정 세션 유지/재로그인 정책
    - MY는 셀러센터 계정 개설 자체가 아직 미확인 상태이므로, 계정 개설 전에는 호출 불가
    """
    raise NotImplementedError("listing-agent: Mass Upload 자동화는 Phase 2 (현재 미구현, Phase 1로 진행 중).")


def publish_individually(uploaded: list[ListingResult]) -> list[ListingResult]:
    """[Phase 2 - 지금 당장 필요하지 않음] Mass Upload된 상품을 하나씩 개별 Publish 한다.
    Phase 1에서는 사람이 쇼피 셀러센터에서 직접 개별 Publish 버튼을 누른다.

    TODO: Phase 2를 진행하기로 하면, Publish 실패 시 재시도 횟수/사유 로깅,
    master-qa-agent로의 반려 전달 포맷 필요.
    """
    raise NotImplementedError("listing-agent: 개별 Publish 자동화는 Phase 2 (현재 미구현, Phase 1로 진행 중).")
